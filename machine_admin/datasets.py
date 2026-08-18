"""Importação cifrada de bases e criação idempotente de itens."""

from __future__ import annotations

import hashlib
import io
import json
import secrets
import zipfile
import csv
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from machine_admin.config import Settings
from machine_admin.models import (
    Dataset,
    DatasetRecord,
    Job,
    JobEvent,
    JobItem,
    Municipality,
)
from machine_admin.security import SecretCipher, fingerprint_identifier
from services.utils import digits_only


DUPLICATE_POLICIES = frozenset({"reject", "keep_first", "keep_all"})
MAX_DATASET_ROWS = 250_000
MAX_DATASET_COLUMNS = 200
MAX_DATASET_CELLS = 2_000_000
MAX_CELL_CHARACTERS = 32_767
MAX_XLSX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024


def _validate_xlsx_payload(payload: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            entries = archive.infolist()
            if len(entries) > 10_000:
                raise ValueError("O XLSX possui componentes demais.")
            if sum(entry.file_size for entry in entries) > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ValueError("O XLSX descompactado excede o limite de segurança.")
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
        try:
            worksheet = workbook.active
            if worksheet.max_row > MAX_DATASET_ROWS + 1:
                raise ValueError(
                    f"A base pode ter no máximo {MAX_DATASET_ROWS:,} registros."
                )
            if worksheet.max_column > MAX_DATASET_COLUMNS:
                raise ValueError(
                    f"A base pode ter no máximo {MAX_DATASET_COLUMNS} colunas."
                )
            if worksheet.max_row * worksheet.max_column > MAX_DATASET_CELLS:
                raise ValueError(
                    "A base excede o limite de 2.000.000 de células. "
                    "Divida o arquivo em bases menores."
                )
        finally:
            workbook.close()
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("O arquivo XLSX está corrompido ou é inválido.") from exc


def _read_table(filename: str, payload: bytes) -> pd.DataFrame:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Formato não aceito. Envie XLSX ou CSV.")
    if suffix == ".csv":
        estimated_rows = payload.count(b"\n") + 1
        if estimated_rows > MAX_DATASET_ROWS + 1:
            raise ValueError(
                f"A base pode ter no máximo {MAX_DATASET_ROWS:,} registros."
            )
        try:
            first_line = payload.splitlines()[0].decode("utf-8-sig")
            estimated_columns = len(next(csv.reader([first_line])))
            if estimated_columns > MAX_DATASET_COLUMNS:
                raise ValueError(
                    f"A base pode ter no máximo {MAX_DATASET_COLUMNS} colunas."
                )
            if estimated_rows * max(estimated_columns, 1) > MAX_DATASET_CELLS:
                raise ValueError(
                    "A base excede o limite de 2.000.000 de células. "
                    "Divida o arquivo em bases menores."
                )
            return pd.read_csv(io.BytesIO(payload), dtype=str)
        except ValueError as exc:
            if "A base " in str(exc):
                raise
            raise ValueError(
                "Não foi possível ler o CSV. Confirme a codificação e o formato do arquivo."
            ) from exc
        except Exception as exc:
            raise ValueError(
                "Não foi possível ler o CSV. Confirme a codificação e o formato do arquivo."
            ) from exc
    _validate_xlsx_payload(payload)
    try:
        return pd.read_excel(io.BytesIO(payload), dtype=str)
    except Exception as exc:
        raise ValueError(
            "Não foi possível ler o XLSX. Confirme se o arquivo não está corrompido."
        ) from exc


def _validate_required_columns(columns: list[str]) -> tuple[str, str | None]:
    """Exige CPF como primeira coluna; matrícula é opcional na segunda."""
    normalized = [column.strip().lstrip("\ufeff").upper() for column in columns]
    if not normalized or normalized[0] != "CPF":
        raise ValueError("Formato inválido: a primeira coluna deve ser CPF.")
    registration_column = (
        columns[1] if len(columns) > 1 and normalized[1] == "MATRICULA" else None
    )
    return columns[0], registration_column


def normalize_cpf(value: object) -> str | None:
    """Normaliza CPF e valida seus dois dígitos verificadores.

    Arquivos Excel frequentemente removem zeros iniciais de colunas numéricas;
    por isso valores com 8 a 10 dígitos continuam recebendo ``zfill`` antes da
    validação oficial.
    """
    digits = digits_only(value)
    if 8 <= len(digits) <= 10:
        digits = digits.zfill(11)
    if len(digits) != 11 or len(set(digits)) == 1:
        return None

    numbers = [int(char) for char in digits]

    def verifier(values: list[int], start_weight: int) -> int:
        remainder = sum(
            number * weight
            for number, weight in zip(values, range(start_weight, 1, -1))
        ) % 11
        return 0 if remainder < 2 else 11 - remainder

    first = verifier(numbers[:9], 10)
    second = verifier(numbers[:9] + [first], 11)
    return digits if numbers[9:] == [first, second] else None


def normalize_duplicate_policy(value: str | None) -> str:
    policy = (value or "keep_first").strip().lower()
    if policy not in DUPLICATE_POLICIES:
        raise ValueError(
            "Política de duplicados inválida. Use reject, keep_first ou keep_all."
        )
    return policy


def _record_identity(cpf: str, registration: str | None) -> tuple[str, str]:
    return cpf, (registration or "").strip().casefold()


def normalise_custom_columns(value: str | list[str] | None) -> list[str]:
    """Normaliza campos definidos no painel sem alterar o schema SQL."""
    values = value.replace("\n", ",").split(",") if isinstance(value, str) else (value or [])
    columns = list(dict.fromkeys(str(item).strip() for item in values if str(item).strip()))
    if len(columns) > 40:
        raise ValueError("Informe no máximo 40 campos personalizados.")
    if any(len(column) > 120 for column in columns):
        raise ValueError("Cada campo personalizado pode ter no máximo 120 caracteres.")
    return columns


def import_dataset(
    session: Session,
    settings: Settings,
    *,
    municipality_slug: str,
    filename: str,
    payload: bytes,
    uploaded_by_id: int,
    custom_columns: str | list[str] | None = None,
    display_name: str | None = None,
    duplicate_policy: str = "keep_first",
    metadata: dict[str, Any] | None = None,
) -> Dataset:
    if not payload or len(payload) > settings.max_upload_bytes:
        raise ValueError("Arquivo vazio ou acima do limite permitido.")
    dataframe = _read_table(filename, payload)
    dataframe.columns = [str(column).strip() for column in dataframe.columns]
    if len(dataframe) > MAX_DATASET_ROWS:
        raise ValueError(f"A base pode ter no máximo {MAX_DATASET_ROWS:,} registros.")
    if len(dataframe.columns) > MAX_DATASET_COLUMNS:
        raise ValueError(f"A base pode ter no máximo {MAX_DATASET_COLUMNS} colunas.")
    if len(dataframe) * max(len(dataframe.columns), 1) > MAX_DATASET_CELLS:
        raise ValueError(
            "A base excede o limite de 2.000.000 de células. "
            "Divida o arquivo em bases menores."
        )
    if any(len(column) > 120 for column in dataframe.columns):
        raise ValueError("Os nomes das colunas podem ter no máximo 120 caracteres.")
    cpf_column, registration_column = _validate_required_columns(list(dataframe.columns))
    municipality = (
        session.get(Municipality, municipality_slug)
        if hasattr(session, "get")
        else None
    )
    input_schema = municipality.input_schema if municipality else {}
    required_fields = {
        str(field).strip().lower() for field in input_schema.get("required", [])
    }
    duplicate_key = [
        str(field).strip().lower()
        for field in input_schema.get(
            "deduplication_key", ["cpf", "registration"]
        )
    ]
    if "registration" in required_fields and registration_column is None:
        raise ValueError(
            "Formato inválido para este convênio: a segunda coluna deve ser MATRICULA."
        )
    extra_columns = normalise_custom_columns(custom_columns)
    policy = normalize_duplicate_policy(duplicate_policy)
    friendly_name = (display_name or Path(filename).stem).strip()
    if not friendly_name:
        raise ValueError("O nome amigável da base é obrigatório.")
    if len(friendly_name) > 160:
        raise ValueError("O nome amigável da base pode ter no máximo 160 caracteres.")
    digest = hashlib.sha256(payload).hexdigest()
    dataset = Dataset(
        municipality_slug=municipality_slug,
        uploaded_by_id=uploaded_by_id,
        original_filename=Path(filename).name,
        display_name=friendly_name,
        storage_path="pending",
        sha256=digest,
        row_count=0,
        duplicate_policy=policy,
        metadata_json=dict(metadata or {}),
        custom_columns=extra_columns,
        status="uploading",
    )
    session.add(dataset)
    session.flush()

    cipher = SecretCipher(settings.master_key)
    directory = settings.storage_dir / "datasets" / str(dataset.id)
    directory.mkdir(parents=True, exist_ok=True)
    encrypted_path = directory / f"{digest}.enc"
    temporary_path = directory / f".{digest}.tmp"
    encrypted_file = cipher.encrypt_bytes(payload, context=f"dataset:{dataset.id}:file")
    temporary_path.write_bytes(encrypted_file)
    temporary_path.replace(encrypted_path)
    dataset.storage_path = str(encrypted_path)

    try:
        records: list[DatasetRecord] = []
        valid_count = 0
        error_count = 0
        error_samples: list[int] = []
        missing_required_count = 0
        missing_required_samples: list[int] = []
        duplicate_count = 0
        duplicate_samples: list[int] = []
        first_row_by_identity: dict[tuple[str, str], int] = {}
        for offset, (_, row) in enumerate(dataframe.iterrows(), start=2):
            digits = normalize_cpf(row.get(cpf_column))
            if digits is None:
                error_count += 1
                if len(error_samples) < 10:
                    error_samples.append(offset)
                continue
            registration = (
                str(row.get(registration_column)).strip()
                if registration_column and not pd.isna(row.get(registration_column))
                else None
            )
            if "registration" in required_fields and not registration:
                missing_required_count += 1
                if len(missing_required_samples) < 10:
                    missing_required_samples.append(offset)
                continue
            identity = _record_identity(
                digits, registration if "registration" in duplicate_key else None
            )
            first_row = first_row_by_identity.get(identity)
            if first_row is not None:
                duplicate_count += 1
                if len(duplicate_samples) < 10:
                    duplicate_samples.append(offset)
                if policy == "reject":
                    raise ValueError(
                        "Base rejeitada: registro duplicado nas linhas "
                        f"{first_row} e {offset}."
                    )
                if policy == "keep_first":
                    continue
            else:
                first_row_by_identity[identity] = offset
            context_id = secrets.token_hex(16)
            raw_row = {
                str(key): (None if pd.isna(value) else str(value))
                for key, value in row.to_dict().items()
            }
            if any(
                len(value) > MAX_CELL_CHARACTERS
                for value in raw_row.values()
                if isinstance(value, str)
            ):
                raise ValueError(
                    f"A linha {offset} contém uma célula acima de "
                    f"{MAX_CELL_CHARACTERS} caracteres."
                )
            for column in extra_columns:
                raw_row.setdefault(column, None)
            records.append(
                DatasetRecord(
                    dataset_id=dataset.id,
                    row_number=offset,
                    encryption_context=context_id,
                    cpf_ciphertext=cipher.encrypt(
                        digits, context=f"record:{context_id}:cpf"
                    ),
                    cpf_fingerprint=fingerprint_identifier(settings.master_key, digits),
                    cpf_last4=digits[-4:],
                    registration=registration,
                    source_ciphertext=cipher.encrypt(
                        json.dumps(raw_row, ensure_ascii=False),
                        context=f"record:{context_id}:source",
                    ),
                    source_data={"columns": list(raw_row)},
                )
            )
            valid_count += 1
            if len(records) >= 1_000:
                session.add_all(records)
                session.flush()
                records.clear()
        if not valid_count:
            raise ValueError("A base não possui registros válidos.")

        if records:
            session.add_all(records)
            session.flush()
        dataset.row_count = valid_count
        dataset.status = "ready"
        dataset.metadata_json = {
            **dataset.metadata_json,
            "import_version": 2,
            "source_columns": list(dataframe.columns),
            "cpf_validation": "checksum",
            "duplicate_key": duplicate_key,
            "invalid_row_count": error_count,
            "missing_required_row_count": missing_required_count,
            "duplicate_row_count": duplicate_count,
        }
        warnings: list[str] = []
        if error_count:
            warnings.append(
                f"{error_count} linha(s) ignorada(s) por CPF inválido; primeiras linhas: "
                + ", ".join(map(str, error_samples))
            )
        if missing_required_count:
            warnings.append(
                f"{missing_required_count} linha(s) ignorada(s) por matrícula ausente; "
                "primeiras linhas: "
                + ", ".join(map(str, missing_required_samples))
            )
        if duplicate_count and policy == "keep_first":
            warnings.append(
                f"{duplicate_count} duplicata(s) ignorada(s) pela política keep_first; "
                "primeiras linhas: " + ", ".join(map(str, duplicate_samples))
            )
        dataset.error_message = " ".join(warnings) or None
        session.flush()
        return dataset
    except Exception:
        temporary_path.unlink(missing_ok=True)
        encrypted_path.unlink(missing_ok=True)
        raise


def delete_dataset_blob(storage_path: str | None) -> None:
    """Remove somente o blob cifrado criado por uma transação abortada."""
    if storage_path and storage_path != "pending":
        Path(storage_path).unlink(missing_ok=True)


def create_job_for_dataset(
    session: Session,
    *,
    dataset: Dataset,
    requested_by_id: int,
) -> Job:
    job = Job(
        municipality_slug=dataset.municipality_slug,
        dataset_id=dataset.id,
        requested_by_id=requested_by_id,
        status="queued",
        total_items=dataset.row_count,
    )
    session.add(job)
    session.flush()
    record_ids = session.scalars(
        select(DatasetRecord.id)
        .where(DatasetRecord.dataset_id == dataset.id)
        .order_by(DatasetRecord.id)
    )
    session.add_all(
        [JobItem(job_id=job.id, dataset_record_id=record_id) for record_id in record_ids]
    )
    session.add(
        JobEvent(
            job_id=job.id,
            event_type="dataset_attached",
            message=f"Base {dataset.id} vinculada com {dataset.row_count} itens.",
        )
    )
    session.flush()
    return job


def attach_dataset_to_job(session: Session, *, job: Job, dataset: Dataset) -> Job:
    if job.status != "awaiting_dataset":
        raise ValueError("O job não está aguardando uma base.")
    if job.municipality_slug != dataset.municipality_slug:
        raise ValueError("A base pertence a outro convênio.")
    job.dataset_id = dataset.id
    job.status = "queued"
    job.total_items = dataset.row_count
    record_ids = session.scalars(
        select(DatasetRecord.id).where(DatasetRecord.dataset_id == dataset.id)
    )
    session.add_all(
        [JobItem(job_id=job.id, dataset_record_id=record_id) for record_id in record_ids]
    )
    session.add(
        JobEvent(
            job_id=job.id,
            event_type="dataset_attached",
            message=f"Base {dataset.id} anexada pelo painel.",
        )
    )
    session.flush()
    return job
